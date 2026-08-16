"""
Import cycle tyre production data from Excel as PRODUCTION ENTRIES (month-by-month).
- Clears ALL existing stock to 0
- Clears ALL existing CycleTyreEntry records
- Clears ALL existing CycleTyreDailyManualEntry records
- Creates production entries for April, May, June, July 2026
- Stock is updated through the entries (1st grade -> stock, 2nd grade -> second_stock)
- Rejected does NOT go to rfm_stock
"""

import os, sys, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'radhu.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

import openpyxl
import datetime
from django.db import transaction
from cycletyres.models import CycleTyreItem, CycleTyreEntry, CycleTyreDailyManualEntry

EXCEL_PATH = r'C:\Users\risha\Desktop\cycle tyre  stock.xlsx'

# Month dates - using 15th of each month as representative date for monthly production
MONTHS = {
    'april':  datetime.date(2026, 4, 15),
    'may':    datetime.date(2026, 5, 15),
    'june':   datetime.date(2026, 6, 15),
    'july':   datetime.date(2026, 7, 15),
}

# Column mapping (0-indexed from iter_rows):
#   Col 0: ply number (6, 7, 8, etc.) or None
#   Col 1: "ply" or None
#   Col 2: size
#   Col 3: material (CTC/NYL)
#   Col 4: brand
#   APRIL:  Col 5=Black, Col 6=B(2nd), Col 7=R(rejected), Col 8=ALL CURING
#   MAY:    Col 9=Black, Col 10=B(2nd), Col 11=R(rejected), Col 12=ALL CURING
#   JUNE:   Col 13=Black, Col 14=B(2nd), Col 15=Rejected, Col 16=ALL CURING
#   JULY:   Col 17=Black, Col 18=B(2nd), Col 19=R(rejected), Col 20=ALL CURING

MONTH_COLS = {
    'april': {'black': 5, 'second': 6, 'rejected': 7, 'all_curing': 8},
    'may':   {'black': 9, 'second': 10, 'rejected': 11, 'all_curing': 12},
    'june':  {'black': 13, 'second': 14, 'rejected': 15, 'all_curing': 16},
    'july':  {'black': 17, 'second': 18, 'rejected': 19, 'all_curing': 20},
}


def safe_int(v):
    if v is None:
        return 0
    try:
        return int(v)
    except (ValueError, TypeError):
        return 0


def find_item_exact(size, mat_raw, brand, p_num):
    size = (size or '').strip()
    mat_raw = (mat_raw or '').strip()
    brand = (brand or '').strip()

    if not size or not mat_raw:
        return None

    if p_num and p_num > 0:
        mat_with_ply = f'{mat_raw} ({p_num} Ply)'
        box_with_ply = f'{p_num} ply'

        # 1. Try material='CTC (6 Ply)'
        it = CycleTyreItem.objects.filter(size__iexact=size, material__iexact=mat_with_ply, brand__iexact=brand).first()
        if it: return it

        # 2. Try box_type='6 ply'
        it = CycleTyreItem.objects.filter(size__iexact=size, box_type__iexact=box_with_ply, brand__iexact=brand).first()
        if it: return it

    # 3. Standard lookup without ply constraint
    it = CycleTyreItem.objects.filter(size__iexact=size, material__iexact=mat_raw, brand__iexact=brand).first()
    if it: return it

    # 4. Fallback search without brand if brand is None/empty
    it = CycleTyreItem.objects.filter(size__iexact=size, material__iexact=mat_raw).first()
    return it


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    data_rows = []

    # Rows 5 to 70 in Excel contain the 66 cycle tyre items
    for r in range(5, 71):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 22)]

        p_num = safe_int(vals[0]) if vals[0] is not None else None
        size = vals[2]
        mat = vals[3]
        brand = vals[4]

        if not size or not mat:
            continue

        row_data = {
            'row_num': r,
            'ply_num': p_num,
            'size': str(size).strip(),
            'material': str(mat).strip(),
            'brand': str(brand).strip() if brand else '',
            'months': {}
        }

        for month_name, cols in MONTH_COLS.items():
            black = safe_int(vals[cols['black']])
            second = safe_int(vals[cols['second']])
            rejected = safe_int(vals[cols['rejected']])
            all_curing = safe_int(vals[cols['all_curing']])

            row_data['months'][month_name] = {
                'all_curing': all_curing,
                'first_grade': black,
                'second_grade': second,
                'rejected_grade': rejected,
            }

        data_rows.append(row_data)

    print(f"Parsed {len(data_rows)} items from Excel")
    print()

    with transaction.atomic():
        # Step 1: Clear all existing entries
        entry_count = CycleTyreEntry.objects.count()
        CycleTyreEntry.objects.all().delete()
        print(f"Deleted {entry_count} existing CycleTyreEntry records")

        manual_count = CycleTyreDailyManualEntry.objects.count()
        CycleTyreDailyManualEntry.objects.all().delete()
        print(f"Deleted {manual_count} existing CycleTyreDailyManualEntry records")

        # Step 2: Reset all stock to 0
        CycleTyreItem.objects.all().update(stock=0, second_stock=0, rfm_stock=0)
        print("Reset all CycleTyreItem stock to 0")
        print()

        # Step 3: Create production entries month by month
        created = 0
        skipped = 0
        matched_items = set()

        for row_data in data_rows:
            item = find_item_exact(
                row_data['size'],
                row_data['material'],
                row_data['brand'],
                row_data['ply_num']
            )

            if not item:
                print(f"ERROR: Item not found for row {row_data['row_num']}: {row_data}")
                continue

            matched_items.add(item.id)

            for month_name in ['april', 'may', 'june', 'july']:
                m = row_data['months'][month_name]
                all_curing = m['all_curing']

                if all_curing <= 0:
                    skipped += 1
                    continue

                first_grade = m['first_grade']
                second_grade = m['second_grade']
                rejected_grade = m['rejected_grade']

                # Sanity check breakdown
                total_breakdown = first_grade + second_grade + rejected_grade
                if total_breakdown != all_curing:
                    first_grade = all_curing - second_grade - rejected_grade
                    if first_grade < 0:
                        first_grade = 0

                date_val = MONTHS[month_name]

                CycleTyreEntry.objects.create(
                    tyre_item=item,
                    entry_type='production',
                    bucket='stock',
                    quantity=first_grade,
                    all_curing=all_curing,
                    first_grade=first_grade,
                    second_grade=second_grade,
                    rejected_grade=rejected_grade,
                    date=date_val,
                    remark=f'{month_name.title()} 2026 production',
                )

                item.stock += first_grade
                item.second_stock += second_grade
                created += 1

            item.save(update_fields=['stock', 'second_stock'])

        print(f"Created {created} production entries")
        print(f"Skipped {skipped} zero-quantity months")
        print(f"Matched {len(matched_items)} unique items")
        print()

        # Print final stock summary
        items = CycleTyreItem.objects.filter(is_active=True)
        total_stock = sum(i.stock for i in items)
        total_second = sum(i.second_stock for i in items)
        total_rfm = sum(i.rfm_stock for i in items)

        print("=" * 70)
        print("FINAL STOCK SUMMARY:")
        print(f"  1st Grade (Black) Stock: {total_stock:,}")
        print(f"  2nd Grade Stock:         {total_second:,}")
        print(f"  R.F.M. Stock:            {total_rfm:,}")
        print(f"  Total Combined:          {total_stock + total_second + total_rfm:,}")
        print()

        for month_name, date_val in MONTHS.items():
            entries = CycleTyreEntry.objects.filter(date=date_val, entry_type='production')
            count = entries.count()
            total_curing = sum(e.all_curing for e in entries)
            total_1st = sum(e.first_grade for e in entries)
            total_2nd = sum(e.second_grade for e in entries)
            total_rej = sum(e.rejected_grade for e in entries)
            print(f"  {month_name.title():>5} 2026 ({date_val}): {count:>2} entries | "
                  f"Curing: {total_curing:>7,} | 1st Grade: {total_1st:>7,} | "
                  f"2nd Grade: {total_2nd:>5,} | Rejected: {total_rej:>5,}")

        print("=" * 70)
        print("DONE!")


if __name__ == '__main__':
    main()
