"""Parse Tally mapping Excel sheets and match rows to stock items."""
from __future__ import annotations

import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from stock.models import TyreItem
from cycletube.models import CycleTubeItem
from cycletyres.models import CycleTyreItem
from .models import TallyItemMapping, MODULE_CHOICES

VALID_MODULES = {choice[0] for choice in MODULE_CHOICES}

HEADER_ALIASES = {
    "tally_item_name": {
        "tally item name",
        "tally item name (from tally)",
        "tally name",
        "item name",
    },
    "module": {
        "target module (tyre / tube / cycletyre)",
        "module (tyre / tube / cycletyre)",
        "module",
        "target module",
    },
    "size": {
        "target item size",
        "target item size / name",
        "size",
        "tyre",
    },
    "pattern": {
        "target pattern / box type / tube type",
        "pattern / box type",
        "pattern",
        "box type",
        "type / box",
        "ply / bsw",
    },
    "material": {
        "target material / quality",
        "material / quality",
        "material",
        "type (tl/tt)",
        "quality",
    },
    "brand": {
        "target brand",
        "brand",
    },
    "item_id": {"item_id", "item id", "id", "target item id"},
    "status": {"status"},
    "notes": {"rule applied / notes", "remarks / notes", "notes", "remarks"},
}

BLANK = {"", "-", "—", "na", "n/a", "none", "null"}

PATTERN_ALIASES = {
    "SUPER": {"SUPER", "RSR", "RSRSUPER", "SUPERRSR"},
    "RSR": {"SUPER", "RSR", "RSRSUPER", "SUPERRSR"},
    "7STAR": {"7STAR", "7START", "7STARS", "SEVENSTAR"},
    "ATTACK": {"ATTACK", "TLATTACK"},
    "BSW": {"BSW", "7PLYBSW", "PLYBSW"},
}


def _norm_key(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _blank(value):
    return str(value or "").strip().lower() in BLANK


def normalize_token(value):
    text = str(value or "").upper()
    text = text.replace("&", " AND ")
    text = re.sub(r"[()\[\]{}]", " ", text)
    text = re.sub(r"[^A-Z0-9]+", "", text)
    return text


def normalize_size(value):
    text = str(value or "").upper().strip()
    text = text.replace("×", "X").replace("*", "X")
    text = re.sub(r"\s+", "", text)
    m = re.match(r"^(\d{2,3})/(\d{2,3})[/](\d{2})$", text)
    if m:
        return f"{m.group(1)}/{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{2,3})/(\d{2,3})-(\d{2})$", text)
    if m:
        return f"{m.group(1)}/{m.group(2)}-{m.group(3)}"
    return text


def _alias_set(token):
    if not token:
        return set()
    extra = set()
    for group in PATTERN_ALIASES.values():
        if token in group:
            extra |= group
    return {token} | extra


def _header_map(header_row):
    mapping = {}
    for idx, cell in enumerate(header_row):
        key = _norm_key(cell)
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _cell(row, header_map, field):
    idx = header_map.get(field)
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    return str(value).strip()


def load_catalog():
    tyres = list(TyreItem.objects.all())
    tubes = list(CycleTubeItem.objects.all())
    cycletyres = list(CycleTyreItem.objects.all())
    return {
        "tyre": tyres,
        "tube": tubes,
        "cycletyre": cycletyres,
    }


def _score_tyre(item, size, pattern, material, tally_name):
    item_size = normalize_size(item.tyre)
    want_size = normalize_size(size)
    if want_size and item_size != want_size:
        return 0
    score = 10 if want_size else 1

    want_pattern = normalize_token(pattern)
    item_pattern = normalize_token(item.pattern)
    if want_pattern:
        if item_pattern in _alias_set(want_pattern) or want_pattern in _alias_set(item_pattern):
            score += 8
        elif want_pattern in item_pattern or item_pattern in want_pattern:
            score += 5
        else:
            return 0

    want_type = normalize_token(material)
    if want_type in {"TL", "TT", "TUBELESS", "TUBE"}:
        want_type = "TL" if want_type in {"TL", "TUBELESS"} else "TT"
        if normalize_token(item.type) == want_type:
            score += 4
        else:
            score -= 2
    else:
        name = (tally_name or "").upper()
        inferred = "TL" if re.search(r"\bTL\b|TUBELESS", name) else ""
        if inferred and normalize_token(item.type) == inferred:
            score += 3
        elif not inferred and normalize_token(item.type) == "TT":
            score += 1
    return score


def _score_cycletyre(item, size, pattern, material, brand):
    item_size = normalize_size(item.size)
    want_size = normalize_size(size)
    if want_size and item_size != want_size and item_size.replace("X", "") != want_size.replace("X", ""):
        # 26x1.5 vs 26X1.5 already normalized; also allow 26x1.5 vs 26 x 1.5
        if normalize_size(item.size.replace(" ", "")) != want_size:
            return 0
    score = 10 if want_size else 1

    want_box = normalize_token(pattern)
    item_box = normalize_token(item.box_type)
    if want_box:
        ply_want = re.search(r"(\d+)(?:PLY|P)", want_box)
        ply_item = re.search(r"(\d+)(?:PLY|P)", item_box)
        if item_box in _alias_set(want_box) or want_box in item_box or item_box in want_box:
            score += 6
        elif ply_want and ply_item and ply_want.group(1) == ply_item.group(1):
            score += 5
        else:
            score -= 3

    want_mat = normalize_token(material)
    if want_mat and want_mat not in {"RUBBER"}:
        if normalize_token(item.material) == want_mat:
            score += 3
        elif want_mat in normalize_token(item.material):
            score += 1

    want_brand = normalize_token(brand)
    item_brand = normalize_token(item.brand)
    if want_brand:
        if item_brand in _alias_set(want_brand) or want_brand in _alias_set(item_brand):
            score += 8
        elif want_brand in item_brand or item_brand in want_brand:
            score += 4
        else:
            return 0
    return score


def _score_tube(item, size, pattern, brand):
    item_size = normalize_size(item.size)
    want_size = normalize_size(size)
    if want_size and item_size != want_size:
        return 0
    score = 10 if want_size else 1

    want_type = normalize_token(pattern)
    item_type = normalize_token(item.type)
    if want_type:
        if item_type == want_type or want_type in item_type or item_type in want_type:
            score += 5
        else:
            score -= 2

    want_brand = normalize_token(brand)
    item_brand = normalize_token(item.brand)
    if want_brand:
        if item_brand in _alias_set(want_brand) or want_brand in _alias_set(item_brand):
            score += 8
        elif want_brand in item_brand or item_brand in want_brand:
            score += 4
        else:
            return 0
    return score


def match_item(catalog, module, size, pattern, material, brand, item_id=None, tally_name=""):
    if item_id:
        try:
            pk = int(item_id)
        except (TypeError, ValueError):
            pk = None
        if pk:
            if module == "tyre":
                found = next((x for x in catalog["tyre"] if x.id == pk), None)
            elif module == "tube":
                found = next((x for x in catalog["tube"] if x.id == pk), None)
            elif module == "cycletyre":
                found = next((x for x in catalog["cycletyre"] if x.id == pk), None)
            else:
                found = None
            if found:
                return found, "matched_by_id"

    items = catalog.get(module) or []
    scored = []
    for item in items:
        if module == "tyre":
            score = _score_tyre(item, size, pattern, material, tally_name)
        elif module == "cycletyre":
            score = _score_cycletyre(item, size, pattern, material, brand)
        elif module == "tube":
            score = _score_tube(item, size, pattern, brand)
        else:
            score = 0
        if score > 0:
            scored.append((score, item))

    if not scored:
        return None, "no_matching_stock_item"

    scored.sort(key=lambda pair: pair[0], reverse=True)
    best_score, best = scored[0]
    ties = [item for score, item in scored if score == best_score]
    if len(ties) > 1 and best_score < 18:
        return None, f"ambiguous_match ({len(ties)} items, score {best_score})"
    return best, f"matched_score_{best_score}"


def parse_mapping_workbook(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    sheet_name = None
    for name in wb.sheetnames:
        if "tally item mapping" in name.lower() or name.lower() == "mapping":
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    header_map = _header_map(header)
    if "tally_item_name" not in header_map:
        # fall back to column B (index 1) used by the master sheet
        header_map["tally_item_name"] = 1
        header_map.setdefault("module", 2)
        header_map.setdefault("size", 3)
        header_map.setdefault("pattern", 4)
        header_map.setdefault("material", 5)
        header_map.setdefault("brand", 6)
        header_map.setdefault("notes", 7)
        header_map.setdefault("status", 8)

    parsed = []
    for row in rows_iter:
        if not row:
            continue
        name = _cell(row, header_map, "tally_item_name")
        if not name:
            continue
        module = _cell(row, header_map, "module").lower().replace(" ", "")
        if module in {"autotyre", "auto"}:
            module = "tyre"
        elif module in {"cycletube", "cy-tube", "cytube"}:
            module = "tube"
        elif module in {"cycle tyre", "cytyre", "cycle-tyre"}:
            module = "cycletyre"

        parsed.append({
            "tally_item_name": name,
            "module": module,
            "size": _cell(row, header_map, "size"),
            "pattern": _cell(row, header_map, "pattern"),
            "material": _cell(row, header_map, "material"),
            "brand": _cell(row, header_map, "brand"),
            "item_id": _cell(row, header_map, "item_id"),
            "status": _cell(row, header_map, "status"),
            "notes": _cell(row, header_map, "notes"),
        })
    wb.close()
    return parsed


def import_mapping_rows(rows, dry_run=False):
    catalog = load_catalog()
    created = updated = skipped = 0
    unmatched = []
    imported_names = []

    for row in rows:
        name = row["tally_item_name"].strip()
        status = (row.get("status") or "").strip().lower()
        module = (row.get("module") or "").strip().lower()

        if status in {"pending review", "review", "skip", "uncertain"} or module in {"review", "pending"}:
            skipped += 1
            unmatched.append({
                "tally_item_name": name,
                "module": module,
                "reason": "skipped_pending_review",
                "notes": row.get("notes") or "",
            })
            continue

        if module not in VALID_MODULES:
            unmatched.append({
                "tally_item_name": name,
                "module": module,
                "reason": "invalid_module (use tyre / tube / cycletyre)",
                "notes": row.get("notes") or "",
            })
            continue

        item, reason = match_item(
            catalog,
            module,
            row.get("size"),
            row.get("pattern"),
            row.get("material"),
            row.get("brand"),
            row.get("item_id"),
            tally_name=name,
        )
        if item is None:
            unmatched.append({
                "tally_item_name": name,
                "module": module,
                "reason": reason,
                "size": row.get("size") or "",
                "pattern": row.get("pattern") or "",
                "brand": row.get("brand") or "",
                "notes": row.get("notes") or "",
            })
            continue

        if dry_run:
            created += 1
            imported_names.append(name)
            continue

        _, was_created = TallyItemMapping.objects.update_or_create(
            tally_item_name=name,
            defaults={"module": module, "item_id": item.id},
        )
        if was_created:
            created += 1
        else:
            updated += 1
        imported_names.append(name)

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "unmatched": unmatched,
        "imported_names": imported_names,
        "imported_count": created + updated,
    }


EXPORT_HEADERS = [
    "S.No",
    "Tally Item Name",
    "Target Module (tyre / tube / cycletyre)",
    "Target Item Size",
    "Target Pattern / Box Type / Tube Type",
    "Target Material / Quality",
    "Target Brand",
    "Rule Applied / Notes",
    "Status",
    "Resolved Item",
    "Item ID",
]


def build_export_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tally Item Mapping"
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(EXPORT_HEADERS)
    for col in range(1, len(EXPORT_HEADERS) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    mappings = TallyItemMapping.objects.all()
    for idx, mapping in enumerate(mappings, start=1):
        item = mapping.get_item()
        size = pattern = material = brand = ""
        if mapping.module == "tyre" and item:
            size, pattern, material = item.tyre, item.pattern, item.type
        elif mapping.module == "cycletyre" and item:
            size, pattern, material, brand = item.size, item.box_type, item.material, item.brand
        elif mapping.module == "tube" and item:
            size, pattern, brand = item.size, item.type, item.brand
        ws.append([
            idx,
            mapping.tally_item_name,
            mapping.module,
            size,
            pattern,
            material,
            brand,
            "",
            "Mapped" if item else "Broken",
            str(item) if item else f"(deleted #{mapping.item_id})",
            mapping.item_id,
        ])

    for col in range(1, len(EXPORT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28

    ref_tyre = wb.create_sheet("Ref - Auto Tyre")
    ref_tyre.append(["ID", "Size", "Pattern", "Type (TL/TT)", "Full Item Name"])
    for item in TyreItem.objects.all():
        ref_tyre.append([item.id, item.tyre, item.pattern, item.type, str(item)])

    ref_ct = wb.create_sheet("Ref - Cycle Tyre")
    ref_ct.append(["ID", "Size", "Ply / BSW", "Material", "Brand", "Full Item Name"])
    for item in CycleTyreItem.objects.all():
        ref_ct.append([item.id, item.size, item.box_type, item.material, item.brand, str(item)])

    ref_tube = wb.create_sheet("Ref - Cycle Tube")
    ref_tube.append(["ID", "Size", "Type / Box", "Brand", "Full Item Name"])
    for item in CycleTubeItem.objects.all():
        ref_tube.append([item.id, item.size, item.type, item.brand, str(item)])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
