import datetime
from django.db import transaction
from django.db.models import Sum, Q
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl

from .models import TyreItem, DailyEntry, DailyProductionManualEntry
from .serializers import TyreItemSerializer, DailyEntrySerializer, DailyProductionManualEntrySerializer

def _recalculate_auto_stocks():
    for item in TyreItem.objects.all():
        p1 = DailyEntry.objects.filter(tyre_item=item, entry_type="production", bucket="stock").aggregate(t=Sum("quantity"))["t"] or 0
        s1 = DailyEntry.objects.filter(tyre_item=item, entry_type="dispatch", bucket="stock").aggregate(t=Sum("quantity"))["t"] or 0
        a1 = DailyEntry.objects.filter(tyre_item=item, entry_type="adjustment", bucket="stock").aggregate(t=Sum("quantity"))["t"] or 0

        p2 = DailyEntry.objects.filter(tyre_item=item, entry_type="production").aggregate(t=Sum("second_grade"))["t"] or 0
        s2 = DailyEntry.objects.filter(tyre_item=item, entry_type="dispatch", bucket="second_stock").aggregate(t=Sum("quantity"))["t"] or 0
        a2 = DailyEntry.objects.filter(tyre_item=item, entry_type="adjustment", bucket="second_stock").aggregate(t=Sum("quantity"))["t"] or 0

        p3 = DailyEntry.objects.filter(tyre_item=item, entry_type="production").aggregate(t=Sum("third_grade"))["t"] or 0
        s3 = DailyEntry.objects.filter(tyre_item=item, entry_type="dispatch", bucket="third_stock").aggregate(t=Sum("quantity"))["t"] or 0
        a3 = DailyEntry.objects.filter(tyre_item=item, entry_type="adjustment", bucket="third_stock").aggregate(t=Sum("quantity"))["t"] or 0

        p_rfm = DailyEntry.objects.filter(tyre_item=item, entry_type="production", bucket="rfm_ok_tyre").aggregate(t=Sum("quantity"))["t"] or 0
        s_rfm = DailyEntry.objects.filter(tyre_item=item, entry_type="dispatch", bucket="rfm_ok_tyre").aggregate(t=Sum("quantity"))["t"] or 0
        a_rfm = DailyEntry.objects.filter(tyre_item=item, entry_type="adjustment", bucket="rfm_ok_tyre").aggregate(t=Sum("quantity"))["t"] or 0

        item.stock = p1 - s1 + a1
        item.second_stock = p2 - s2 + a2
        item.third_stock = p3 - s3 + a3
        item.rfm_ok_tyre = p_rfm - s_rfm + a_rfm
        item.save(update_fields=["stock", "second_stock", "third_stock", "rfm_ok_tyre"])


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard(request):
    _recalculate_auto_stocks()

    today = datetime.date.today()
    start_date_param = request.query_params.get("start_date", "").strip()
    end_date_param = request.query_params.get("end_date", "").strip()
    month_param = request.query_params.get("month", "").strip()

    filter_start = None
    filter_end = None
    is_filtered_range = False

    if start_date_param and end_date_param:
        try:
            filter_start = datetime.datetime.strptime(start_date_param, "%Y-%m-%d").date()
            filter_end = datetime.datetime.strptime(end_date_param, "%Y-%m-%d").date()
            is_filtered_range = True
        except ValueError:
            pass

    if not is_filtered_range:
        if not month_param:
            month_param = f"{today.year}-{today.month:02d}"

        if month_param != "all":
            try:
                parts = month_param.split("-")
                year, month = int(parts[0]), int(parts[1])
                filter_start = datetime.date(year, month, 1)
                if month == 12:
                    filter_end = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
                else:
                    filter_end = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
                is_filtered_range = True
            except (ValueError, IndexError):
                month_param = "all"
                is_filtered_range = False

    items = TyreItem.objects.filter(is_active=True)
    items_data = []

    tot_prev_first = 0
    tot_prev_second = 0
    tot_prev_third = 0
    tot_prod_total = 0
    tot_prod_first = 0
    tot_prod_second = 0
    tot_prod_third = 0
    tot_sale_first = 0
    tot_sale_second = 0
    tot_sale_third = 0
    tot_rfm = 0
    tot_closing_first = 0
    tot_closing_second = 0
    tot_closing_third = 0
    tot_total_closing = 0

    for item in items:
        if is_filtered_range and filter_start and filter_end:
            # Previous closing (before start date)
            p1_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="production", bucket="stock", date__lt=filter_start
            ).aggregate(t=Sum("quantity"))["t"] or 0
            s1_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="dispatch", bucket="stock", date__lt=filter_start
            ).aggregate(t=Sum("quantity"))["t"] or 0
            a1_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="adjustment", bucket="stock", date__lt=filter_start
            ).aggregate(t=Sum("quantity"))["t"] or 0

            p2_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="production", date__lt=filter_start
            ).aggregate(t=Sum("second_grade"))["t"] or 0
            s2_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="dispatch", bucket="second_stock", date__lt=filter_start
            ).aggregate(t=Sum("quantity"))["t"] or 0
            a2_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="adjustment", bucket="second_stock", date__lt=filter_start
            ).aggregate(t=Sum("quantity"))["t"] or 0

            p3_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="production", date__lt=filter_start
            ).aggregate(t=Sum("third_grade"))["t"] or 0
            s3_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="dispatch", bucket="third_stock", date__lt=filter_start
            ).aggregate(t=Sum("quantity"))["t"] or 0
            a3_prev = DailyEntry.objects.filter(
                tyre_item=item, entry_type="adjustment", bucket="third_stock", date__lt=filter_start
            ).aggregate(t=Sum("quantity"))["t"] or 0

            prev_first = p1_prev - s1_prev + a1_prev
            prev_second = p2_prev - s2_prev + a2_prev
            prev_third = p3_prev - s3_prev + a3_prev

            # In-range production
            range_prod = DailyEntry.objects.filter(
                tyre_item=item, entry_type="production", date__gte=filter_start, date__lte=filter_end
            )
            p1_range = range_prod.filter(bucket="stock").aggregate(t=Sum("quantity"))["t"] or 0
            p2_range = range_prod.aggregate(t=Sum("second_grade"))["t"] or 0
            p3_range = range_prod.aggregate(t=Sum("third_grade"))["t"] or 0

            # In-range sales / dispatches
            s1_range = DailyEntry.objects.filter(
                tyre_item=item, entry_type="dispatch", bucket="stock", date__gte=filter_start, date__lte=filter_end
            ).aggregate(t=Sum("quantity"))["t"] or 0
            s2_range = DailyEntry.objects.filter(
                tyre_item=item, entry_type="dispatch", bucket="second_stock", date__gte=filter_start, date__lte=filter_end
            ).aggregate(t=Sum("quantity"))["t"] or 0
            s3_range = DailyEntry.objects.filter(
                tyre_item=item, entry_type="dispatch", bucket="third_stock", date__gte=filter_start, date__lte=filter_end
            ).aggregate(t=Sum("quantity"))["t"] or 0

            # In-range adjustments
            a1_range = DailyEntry.objects.filter(
                tyre_item=item, entry_type="adjustment", bucket="stock", date__gte=filter_start, date__lte=filter_end
            ).aggregate(t=Sum("quantity"))["t"] or 0
            a2_range = DailyEntry.objects.filter(
                tyre_item=item, entry_type="adjustment", bucket="second_stock", date__gte=filter_start, date__lte=filter_end
            ).aggregate(t=Sum("quantity"))["t"] or 0
            a3_range = DailyEntry.objects.filter(
                tyre_item=item, entry_type="adjustment", bucket="third_stock", date__gte=filter_start, date__lte=filter_end
            ).aggregate(t=Sum("quantity"))["t"] or 0

            closing_first = prev_first + p1_range - s1_range + a1_range
            closing_second = prev_second + p2_range - s2_range + a2_range
            closing_third = prev_third + p3_range - s3_range + a3_range
            rfm = item.rfm_ok_tyre
            total_closing = closing_first + closing_second + closing_third + rfm

            prod_first = p1_range
            prod_second = p2_range
            prod_third = p3_range
            prod_total = p1_range + p2_range + p3_range

            sale_first = s1_range
            sale_second = s2_range
            sale_third = s3_range

        else:
            # All time
            prev_first = 0
            prev_second = 0
            prev_third = 0

            all_prod = DailyEntry.objects.filter(tyre_item=item, entry_type="production")
            prod_first = all_prod.filter(bucket="stock").aggregate(t=Sum("quantity"))["t"] or 0
            prod_second = all_prod.aggregate(t=Sum("second_grade"))["t"] or 0
            prod_third = all_prod.aggregate(t=Sum("third_grade"))["t"] or 0
            prod_total = prod_first + prod_second + prod_third

            sale_first = DailyEntry.objects.filter(tyre_item=item, entry_type="dispatch", bucket="stock").aggregate(t=Sum("quantity"))["t"] or 0
            sale_second = DailyEntry.objects.filter(tyre_item=item, entry_type="dispatch", bucket="second_stock").aggregate(t=Sum("quantity"))["t"] or 0
            sale_third = DailyEntry.objects.filter(tyre_item=item, entry_type="dispatch", bucket="third_stock").aggregate(t=Sum("quantity"))["t"] or 0

            closing_first = item.stock
            closing_second = item.second_stock
            closing_third = item.third_stock
            rfm = item.rfm_ok_tyre
            total_closing = closing_first + closing_second + closing_third + rfm

        serialized = TyreItemSerializer(item).data
        serialized.update({
            "prev_closing_first": prev_first,
            "prev_closing_second": prev_second,
            "prev_closing_third": prev_third,
            "month_prod_total": prod_total,
            "month_prod_first": prod_first,
            "month_prod_second": prod_second,
            "month_prod_third": prod_third,
            "month_sale_first": sale_first,
            "month_sale_second": sale_second,
            "month_sale_third": sale_third,
            "rfm_ok_tyre": rfm,
            "closing_first": closing_first,
            "closing_second": closing_second,
            "closing_third": closing_third,
            "total_closing": total_closing,
        })
        items_data.append(serialized)

        tot_prev_first += prev_first
        tot_prev_second += prev_second
        tot_prev_third += prev_third
        tot_prod_total += prod_total
        tot_prod_first += prod_first
        tot_prod_second += prod_second
        tot_prod_third += prod_third
        tot_sale_first += sale_first
        tot_sale_second += sale_second
        tot_sale_third += sale_third
        tot_rfm += rfm
        tot_closing_first += closing_first
        tot_closing_second += closing_second
        tot_closing_third += closing_third
        tot_total_closing += total_closing

    # Available months list
    entry_dates = DailyEntry.objects.dates("date", "month", order="DESC")
    month_set = set()
    available_months = [{"value": "all", "label": "All Time / Overall"}]

    standard_months = [
        ("2026-08", "August 2026"),
        ("2026-07", "July 2026"),
        ("2026-06", "June 2026"),
        ("2026-05", "May 2026"),
        ("2026-04", "April 2026"),
    ]
    for val, lbl in standard_months:
        available_months.append({"value": val, "label": lbl})
        month_set.add(val)

    for ed in entry_dates:
        val = f"{ed.year}-{ed.month:02d}"
        if val not in month_set:
            lbl = ed.strftime("%B %Y")
            available_months.append({"value": val, "label": lbl})
            month_set.add(val)

    today_prod_all = DailyEntry.objects.filter(entry_type="production", date=today).aggregate(
        q=Sum("quantity"), s=Sum("second_grade"), t=Sum("third_grade")
    )
    today_prod_count = (today_prod_all["q"] or 0) + (today_prod_all["s"] or 0) + (today_prod_all["t"] or 0)
    today_disp_count = DailyEntry.objects.filter(entry_type="dispatch", date=today).aggregate(Sum("quantity"))["quantity__sum"] or 0

    return Response({
        "selected_month": month_param or "custom",
        "start_date": str(filter_start) if filter_start else None,
        "end_date": str(filter_end) if filter_end else None,
        "available_months": available_months,
        "stats": {
            "today_production": today_prod_count,
            "today_dispatch": today_disp_count,
            "month_prod_total": tot_prod_total,
            "month_prod_first": tot_prod_first,
            "month_prod_second": tot_prod_second,
            "month_prod_third": tot_prod_third,
            "month_sale_first": tot_sale_first,
            "closing_first": tot_closing_first,
            "closing_second": tot_closing_second,
            "closing_third": tot_closing_third,
            "total_closing": tot_total_closing,
        },
        "totals": {
            "prev_closing_first": tot_prev_first,
            "prev_closing_second": tot_prev_second,
            "prev_closing_third": tot_prev_third,
            "month_prod_total": tot_prod_total,
            "month_prod_first": tot_prod_first,
            "month_prod_second": tot_prod_second,
            "month_prod_third": tot_prod_third,
            "month_sale_first": tot_sale_first,
            "month_sale_second": tot_sale_second,
            "month_sale_third": tot_sale_third,
            "rfm_ok_tyre": tot_rfm,
            "closing_first": tot_closing_first,
            "closing_second": tot_closing_second,
            "closing_third": tot_closing_third,
            "total_closing": tot_total_closing,
        },
        "items": items_data,
    })

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_tyre(request):
    serializer = TyreItemSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def tyres_list(request):
    items = TyreItem.objects.filter(is_active=True)
    serializer = TyreItemSerializer(items, many=True)
    return Response(serializer.data)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_production(request):
    data = request.data
    tyre_id = data.get("tyre_item")
    try:
        item = TyreItem.objects.get(id=tyre_id)
    except TyreItem.DoesNotExist:
        return Response({"error": "Tyre not found"}, status=status.HTTP_404_NOT_FOUND)

    all_curing = int(data.get("all_curing", 0))
    production_tyre = int(data.get("production_tyre", 0))
    repair = int(data.get("repair", 0))
    second_grade = int(data.get("second_grade", 0))
    third_grade = int(data.get("third_grade", 0))
    lose_tyre = int(data.get("lose_tyre", 0))
    
    packing = (all_curing + repair + production_tyre) - (second_grade + third_grade + lose_tyre)
    if packing < 0:
        return Response({"error": "Packing quantity cannot be negative"}, status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        item.stock += packing
        item.second_stock += second_grade
        item.third_stock += third_grade
        item.save(update_fields=["stock", "second_stock", "third_stock"])

        entry = DailyEntry.objects.create(
            tyre_item=item,
            entry_type="production",
            bucket="stock",
            quantity=packing,
            date=data.get("date"),
            all_curing=all_curing,
            production_tyre=production_tyre,
            repair=repair,
            second_grade=second_grade,
            third_grade=third_grade,
            lose_tyre=lose_tyre,
            actual_weight=data.get("actual_weight") or None,
            remark=data.get("remark", ""),
            user=request.user
        )

    return Response({"message": "Production added successfully", "packing": packing})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def recent_production(request):
    entries = DailyEntry.objects.filter(entry_type="production").select_related("tyre_item", "user")[:15]
    serializer = DailyEntrySerializer(entries, many=True)
    return Response(serializer.data)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_dispatch(request):
    data = request.data
    tyre_id = data.get("tyre_item")
    try:
        item = TyreItem.objects.get(id=tyre_id)
    except TyreItem.DoesNotExist:
        return Response({"error": "Tyre not found"}, status=status.HTTP_404_NOT_FOUND)

    bill_number = data.get("bill_number", "").strip()
    if bill_number:
        if DailyEntry.objects.filter(entry_type="dispatch", bill_number__iexact=bill_number).exists():
            return Response({"error": "Duplicate bill number"}, status=status.HTTP_400_BAD_REQUEST)

    bucket = data.get("bucket", "stock")
    qty = int(data.get("quantity", 0))
    current = getattr(item, bucket)

    if qty > current:
        return Response({"error": f"sirf {current} available hai"}, status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        setattr(item, bucket, current - qty)
        item.save(update_fields=[bucket])

        DailyEntry.objects.create(
            tyre_item=item,
            entry_type="dispatch",
            bucket=bucket,
            quantity=qty,
            date=data.get("date"),
            bill_number=bill_number,
            remark=data.get("remark", ""),
            user=request.user
        )

    return Response({"message": "Dispatch added successfully"})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def recent_dispatch(request):
    entries = DailyEntry.objects.filter(entry_type="dispatch").select_related("tyre_item", "user")[:15]
    serializer = DailyEntrySerializer(entries, many=True)
    return Response(serializer.data)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_adjustment(request):
    data = request.data
    tyre_id = data.get("tyre_item")
    try:
        item = TyreItem.objects.get(id=tyre_id)
    except TyreItem.DoesNotExist:
        return Response({"error": "Tyre not found"}, status=status.HTTP_404_NOT_FOUND)

    bucket = data.get("bucket", "stock")
    action = data.get("action")
    qty = int(data.get("quantity", 0))
    current = getattr(item, bucket)

    if action == "subtract" and qty > current:
        return Response({"error": "Not enough stock for adjustment"}, status=status.HTTP_400_BAD_REQUEST)

    signed_qty = qty if action == "add" else -qty

    with transaction.atomic():
        setattr(item, bucket, current + signed_qty)
        item.save(update_fields=[bucket])

        DailyEntry.objects.create(
            tyre_item=item,
            entry_type="adjustment",
            bucket=bucket,
            quantity=signed_qty,
            date=data.get("date"),
            remark=data.get("remark", ""),
            user=request.user
        )

    return Response({"message": "Adjustment successful"})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def entries_log(request):
    date_param = request.GET.get("date")
    month_param = request.GET.get("month")
    entry_type = request.GET.get("type")

    entries = DailyEntry.objects.all()

    if date_param:
        entries = entries.filter(date=date_param)
    elif month_param:
        try:
            year, month = month_param.split("-")
            entries = entries.filter(date__year=year, date__month=month)
        except:
            pass

    if entry_type:
        entries = entries.filter(entry_type=entry_type)

    entries = entries.select_related("tyre_item", "user")[:500]
    serializer = DailyEntrySerializer(entries, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def monthly_report(request):
    month_param = request.GET.get("month")
    today = datetime.date.today()
    if month_param:
        try:
            year, month = map(int, month_param.split("-"))
            target_date = datetime.date(year, month, 1)
        except:
            target_date = today.replace(day=1)
    else:
        target_date = today.replace(day=1)

    items = TyreItem.objects.filter(is_active=True)
    report_items = []

    total_curing = 0
    total_despatch = 0

    for item in items:
        prod = DailyEntry.objects.filter(tyre_item=item, entry_type="production", date__year=target_date.year, date__month=target_date.month).aggregate(Sum("quantity"))['quantity__sum'] or 0
        disp = DailyEntry.objects.filter(tyre_item=item, entry_type="dispatch", date__year=target_date.year, date__month=target_date.month).aggregate(Sum("quantity"))['quantity__sum'] or 0

        if prod > 0 or disp > 0:
            report_items.append({
                "tyre_id": item.id,
                "tyre_name": item.tyre,
                "pattern": item.pattern,
                "type": item.type,
                "monthly_curing": prod,
                "monthly_despatch": disp
            })
            total_curing += prod
            total_despatch += disp

    grand_total_stock = sum(item.total_stock for item in items)
    net_balance = total_curing - total_despatch

    trend_data = []
    for i in range(5, -1, -1):
        m = target_date.month - i
        y = target_date.year
        if m <= 0:
            m += 12
            y -= 1
        
        prod = DailyEntry.objects.filter(entry_type="production", date__year=y, date__month=m).aggregate(Sum("quantity"))['quantity__sum'] or 0
        disp = DailyEntry.objects.filter(entry_type="dispatch", date__year=y, date__month=m).aggregate(Sum("quantity"))['quantity__sum'] or 0
        
        month_label = datetime.date(y, m, 1).strftime("%b %Y")
        trend_data.append({
            "month": month_label,
            "production": prod,
            "dispatch": disp
        })

    return Response({
        "items": report_items,
        "totals": {
            "total_curing": total_curing,
            "total_despatch": total_despatch,
            "net_balance": net_balance,
            "grand_total_stock": grand_total_stock,
        },
        "trend": trend_data
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def production_sheet(request):
    date_param = request.GET.get("date")
    month_param = request.GET.get("month")
    
    if date_param:
        entries = DailyEntry.objects.filter(entry_type="production", date=date_param).select_related("tyre_item")
    elif month_param:
        try:
            year, month = month_param.split("-")
            entries = DailyEntry.objects.filter(entry_type="production", date__year=year, date__month=month).select_related("tyre_item")
        except:
            today = datetime.date.today()
            entries = DailyEntry.objects.filter(entry_type="production", date__year=today.year, date__month=today.month).select_related("tyre_item")
    else:
        today = datetime.date.today()
        entries = DailyEntry.objects.filter(entry_type="production", date__year=today.year, date__month=today.month).select_related("tyre_item")

    # Assuming no pagination logic for now as it's not strictly required in standard response size and complexity might break exactness without DRF paginator setup. We will return all matching.
    
    data_list = []
    
    totals = {
        "all_curing": 0, "production_tyre": 0, "repair": 0, "second_grade": 0, 
        "third_grade": 0, "lose_tyre": 0, "packing": 0, "rfm_adjustment": 0
    }
    
    labels = []
    curing_data = []
    packing_data = []
    repair_data = []
    second_data = []
    third_data = []
    lose_data = []
    
    if date_param:
        for entry in entries:
            rfm = DailyEntry.objects.filter(entry_type="adjustment", bucket="rfm_ok_tyre", date=entry.date, tyre_item=entry.tyre_item).aggregate(Sum("quantity"))['quantity__sum'] or 0
            data_list.append({
                "tyre_name": str(entry.tyre_item),
                "all_curing": entry.all_curing,
                "production_tyre": entry.production_tyre,
                "repair": entry.repair,
                "second_grade": entry.second_grade,
                "third_grade": entry.third_grade,
                "lose_tyre": entry.lose_tyre,
                "packing": entry.quantity,
                "rfm_adjustment": rfm
            })
            totals["all_curing"] += entry.all_curing
            totals["production_tyre"] += entry.production_tyre
            totals["repair"] += entry.repair
            totals["second_grade"] += entry.second_grade
            totals["third_grade"] += entry.third_grade
            totals["lose_tyre"] += entry.lose_tyre
            totals["packing"] += entry.quantity
            totals["rfm_adjustment"] += rfm
            
            labels.append(str(entry.tyre_item))
            curing_data.append(entry.all_curing)
            packing_data.append(entry.quantity)
            repair_data.append(entry.repair)
            second_data.append(entry.second_grade)
            third_data.append(entry.third_grade)
            lose_data.append(entry.lose_tyre)
            
    else:
        # aggregate by tyre item
        tyre_aggregates = {}
        for entry in entries:
            tid = entry.tyre_item.id
            if tid not in tyre_aggregates:
                tyre_aggregates[tid] = {
                    "tyre_name": str(entry.tyre_item),
                    "all_curing": 0,
                    "production_tyre": 0,
                    "repair": 0,
                    "second_grade": 0,
                    "third_grade": 0,
                    "lose_tyre": 0,
                    "packing": 0,
                    "rfm_adjustment": 0
                }
            
            tyre_aggregates[tid]["all_curing"] += entry.all_curing
            tyre_aggregates[tid]["production_tyre"] += entry.production_tyre
            tyre_aggregates[tid]["repair"] += entry.repair
            tyre_aggregates[tid]["second_grade"] += entry.second_grade
            tyre_aggregates[tid]["third_grade"] += entry.third_grade
            tyre_aggregates[tid]["lose_tyre"] += entry.lose_tyre
            tyre_aggregates[tid]["packing"] += entry.quantity
            
        for tid, agg in tyre_aggregates.items():
            data_list.append(agg)
            for k in totals.keys():
                totals[k] += agg[k]
                
            labels.append(agg["tyre_name"])
            curing_data.append(agg["all_curing"])
            packing_data.append(agg["packing"])
            repair_data.append(agg["repair"])
            second_data.append(agg["second_grade"])
            third_data.append(agg["third_grade"])
            lose_data.append(agg["lose_tyre"])
            
    return Response({
        "data": data_list,
        "totals": totals,
        "chart": {
            "labels": labels,
            "curing": curing_data,
            "values": packing_data,
            "repair": repair_data,
            "second": second_data,
            "third": third_data,
            "lose": lose_data
        }
    })

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def production_sheet_export(request):
    # Simpler version leveraging the same logic, then exporting to openpyxl
    response = production_sheet(request._request)
    if response.status_code != 200:
        return response
        
    data = response.data["data"]
    totals = response.data["totals"]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Sheet"
    
    headers = ["Tyre Item", "All Curing", "Production Tyre", "Repair", "2nd Grade", "3rd Grade", "Lose Tyre", "Packing", "RFM Adjustment"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row["tyre_name"], row["all_curing"], row["production_tyre"], row["repair"],
            row["second_grade"], row["third_grade"], row["lose_tyre"], row["packing"], row["rfm_adjustment"]
        ])
        
    ws.append([
        "TOTAL", totals["all_curing"], totals["production_tyre"], totals["repair"],
        totals["second_grade"], totals["third_grade"], totals["lose_tyre"], totals["packing"], totals["rfm_adjustment"]
    ])
    
    http_response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    http_response["Content-Disposition"] = 'attachment; filename="production_sheet.xlsx"'
    wb.save(http_response)
    return http_response

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def daily_summary(request):
    if request.method == "POST":
        data = request.data
        entry_date = data.get("entry_date")
        if not entry_date:
            return Response({"error": "entry_date is required"}, status=status.HTTP_400_BAD_REQUEST)
            
        obj, created = DailyProductionManualEntry.objects.update_or_create(
            date=entry_date,
            defaults={
                "parchi_kg": data.get("parchi_kg", 0),
                "mixing_actual_compound": data.get("mixing_actual_compound", 0),
                "wastage": data.get("wastage", 0)
            }
        )
        return Response({"message": "Saved successfully", "date": entry_date})

    # GET method
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    
    today = datetime.date.today()
    if not from_date:
        from_date = today.replace(day=1)
    if not to_date:
        to_date = today
        
    entries = DailyEntry.objects.filter(entry_type="production", date__gte=from_date, date__lte=to_date).select_related("tyre_item")
    
    dates_data = {}
    for entry in entries:
        d = str(entry.date)
        if d not in dates_data:
            dates_data[d] = {
                "date": d,
                "curing": 0,
                "packing": 0,
                "theoretical_kg": 0
            }
        dates_data[d]["curing"] += entry.all_curing
        dates_data[d]["packing"] += entry.quantity
        dates_data[d]["theoretical_kg"] += float(entry.all_curing) * float(entry.tyre_item.weight)
        
    manual_entries = DailyProductionManualEntry.objects.filter(date__gte=from_date, date__lte=to_date)
    manual_dict = {str(me.date): me for me in manual_entries}
    
    results = []
    COMPOUND_PERCENT = 0.877
    
    for d, data in dates_data.items():
        me = manual_dict.get(d)
        parchi_kg = float(me.parchi_kg) if me else 0.0
        mixing_actual = float(me.mixing_actual_compound) if me else 0.0
        wastage = float(me.wastage) if me else 0.0
        
        theoretical_kg = data["theoretical_kg"]
        difference = parchi_kg - theoretical_kg  
        theoretical_total_compound = theoretical_kg * COMPOUND_PERCENT
        variance = mixing_actual - theoretical_total_compound
        
        results.append({
            "date": d,
            "curing": data["curing"],
            "packing": data["packing"],
            "theoretical_kg": round(theoretical_kg, 2),
            "parchi_kg": round(parchi_kg, 2),
            "difference": round(difference, 2),
            "theoretical_compound": round(theoretical_total_compound, 2),
            "mixing_actual_compound": round(mixing_actual, 2),
            "variance": round(variance, 2),
            "wastage": round(wastage, 2)
        })
        
    results.sort(key=lambda x: x["date"], reverse=True)
    return Response(results)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_production_excel(request):
    import openpyxl

    file_obj = request.FILES.get("file") or request.FILES.get("excel_file")
    if not file_obj:
        return Response({"error": "No file uploaded. Please select an Excel file."}, status=status.HTTP_400_BAD_REQUEST)

    clear_existing = str(request.data.get("clear_existing", "true")).lower() in ["true", "1", "yes"]

    import_date_str = request.data.get("import_date") or request.data.get("date")
    user_date = None
    if import_date_str:
        try:
            user_date = datetime.datetime.strptime(str(import_date_str).strip(), "%Y-%m-%d").date()
        except ValueError:
            pass

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception as e:
        return Response({"error": f"Failed to open Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    def safe_int(v):
        if v is None: return 0
        try: return int(float(v))
        except: return 0

    data_rows = []
    for r in range(2, ws.max_row + 1):
        tyre_name = ws.cell(row=r, column=1).value or ws.cell(row=r, column=2).value
        if not tyre_name or str(tyre_name).strip().lower() in ["total", "totals", "tyre", "size"]:
            continue

        curing = safe_int(ws.cell(row=r, column=3).value or ws.cell(row=r, column=4).value)
        prod = safe_int(ws.cell(row=r, column=5).value)
        repair = safe_int(ws.cell(row=r, column=6).value)
        second = safe_int(ws.cell(row=r, column=7).value)
        third = safe_int(ws.cell(row=r, column=8).value)
        lose = safe_int(ws.cell(row=r, column=9).value)

        if curing <= 0 and prod <= 0:
            continue

        data_rows.append({
            'row_num': r,
            'name': str(tyre_name).strip(),
            'all_curing': curing,
            'production_tyre': prod,
            'repair': repair,
            'second_grade': second,
            'third_grade': third,
            'lose_tyre': lose,
        })

    if not data_rows:
        return Response({"error": "No valid Auto Tyre production data found in Excel sheet."}, status=status.HTTP_400_BAD_REQUEST)

    entry_date = user_date or datetime.date.today()

    with transaction.atomic():
        if clear_existing:
            DailyEntry.objects.filter(entry_type="production").delete()
            TyreItem.objects.all().update(stock=0, repair_tyre_stock=0, rfm_ok_tyre=0)

        created_count = 0
        total_curing_imported = 0
        total_packing_imported = 0

        for row in data_rows:
            raw_name = row['name']
            item = TyreItem.objects.filter(tyre__iexact=raw_name).first()
            if not item:
                parts = raw_name.split()
                size = parts[0] if parts else raw_name
                pattern = " ".join(parts[1:-1]) if len(parts) > 2 else (parts[1] if len(parts) > 1 else "DEFAULT")
                ttype = parts[-1] if len(parts) > 1 and parts[-1].upper() in ["TL", "TT"] else "TT"
                item, _ = TyreItem.objects.get_or_create(
                    tyre=size, pattern=pattern, type=ttype,
                    defaults={"stock": 0, "is_active": True}
                )

            all_curing = row['all_curing']
            repair = row['repair']
            prod_tyre = row['production_tyre']
            second = row['second_grade']
            third = row['third_grade']
            lose = row['lose_tyre']

            packing = (all_curing + repair + prod_tyre) - (second + third + lose)
            if packing < 0:
                packing = 0

            item.stock += packing
            item.save(update_fields=["stock"])

            DailyEntry.objects.create(
                tyre_item=item,
                entry_type="production",
                bucket="stock",
                quantity=packing,
                date=entry_date,
                all_curing=all_curing,
                production_tyre=prod_tyre,
                repair=repair,
                second_grade=second,
                third_grade=third,
                lose_tyre=lose,
                remark="Excel Bulk Import",
                user=request.user
            )

            created_count += 1
            total_curing_imported += all_curing
            total_packing_imported += packing

    return Response({
        "ok": True,
        "message": f"Successfully imported {created_count} production entries.",
        "created_entries": created_count,
        "total_curing": total_curing_imported,
        "total_packing": total_packing_imported,
        "entry_date": str(entry_date)
    })


